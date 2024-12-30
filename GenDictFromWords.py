import sqlite3
from pathlib import Path, PurePath

from openpyxl import Workbook, load_workbook

from ECDICT import stardict

MYSQLITE = 'ecdictSqlite.db'


def find_lemma(exchange, orig_word):
    """找到ECDICT exchange字段的lemma字符串

    Note:
        ecdict exchange列中0: 代表Lemma，如 perceived 的 Lemma 是 perceive
    """

    last_start_index = exchange.rfind('0:')

    if last_start_index == -1 or last_start_index + 3 > len(exchange):
        return orig_word

    last_end_index = exchange.rfind('/', last_start_index)
    if last_end_index == -1:
        return exchange[last_start_index + 2:]
    else:
        return exchange[last_start_index + 2:last_end_index]


def init_ecdict_sqlite():
    """将ecdict.csv转换为sqlite数据库，并新增lemma列

    """
    stardict.convert_dict(MYSQLITE, 'ECDICT/ecdict.csv')

    con = sqlite3.connect(MYSQLITE)
    cur = con.cursor()
    cur.execute("ALTER TABLE stardict ADD lemma TEXT")

    # 查询所有行的数据
    select_query = "SELECT word, exchange FROM stardict"
    cur.execute(select_query)
    rows = cur.fetchall()

    # 遍历每一行，根据 word和exchange 的值更新新建列lemma
    for row in rows:
        word = row[0]
        exchange = row[1]

        lemma = find_lemma(exchange, word)

        # 更新 lemma 列的值
        update_query = "UPDATE stardict SET lemma = ? WHERE word = ?"
        cur.execute(update_query, (lemma, word))

    con.commit()
    con.close()


def xlsx_write(word_lines, columns, filename, sheetname='sheet1'):
    """将列表文件写入filename文件

    """

    if Path(filename).exists():
        wb = load_workbook(filename)
    else:
        wb = Workbook()

    # 如果sheetname已存在的话，openpyxl会依次在sheetname后追加1,2,3...
    worksheet = wb.create_sheet(sheetname)

    rows = len(word_lines)
    if rows <= 1:
        print(f'word_lines行数为{rows}，有误，不写入文件，推出')
        return

    cols = len(columns)
    if cols != len(word_lines[0]):
        print(
            '参数 columns 有 {cols}列，但word_lines第一行有{len(word_lines[0])}列\n列数匹配有误，程序退出'
        )

    for j in range(cols):
        worksheet.cell(1, j + 1, columns[j])

    # Iterate over the data and write it out row by row.
    for i in range(rows):
        for j in range(cols):
            worksheet.cell(i + 2, j + 1, word_lines[i][j])

    wb.save(filename)
    wb.close()


def write_from_file(input_file):
    """从生字表文件（每单词一行）获取单词，输出包含两个sheet的excel文件，一个sheet含单词、词根、音标和中文释义。另一个sheet由前者词根生成,含单词，音标和中文释义。

    """
    with open(input_file, 'r') as f:
        words = [line.strip() for line in f.readlines()]

    con = sqlite3.connect(MYSQLITE)
    cur = con.cursor()

    # 查询所有行的数据
    holders = ', '.join('?' for _ in words)
    select_query = f"SELECT word, lemma, phonetic, translation FROM stardict WHERE word IN ({holders})"
    cur.execute(select_query, words)
    rows = cur.fetchall()
    con.commit()
    con.close()

    out_file = PurePath(input_file).stem + '_dict.xlsx'

    xlsx_write(rows, ['word', 'lemma', 'phonetic', 'translation'], out_file,
               'Words_Sheet')

    lemmas = [j[1] for j in rows if j[1] is not None]
    write_from_words(lemmas, out_file)


def write_from_words(words, out_file):
    """从单词表（每单词一行）获取单词，输出单词、音标和中文释义

    """
    con = sqlite3.connect(MYSQLITE)
    cur = con.cursor()

    # 查询所有行的数据
    holders = ', '.join('?' for _ in words)
    select_query = f"SELECT word, phonetic, translation FROM stardict WHERE word IN ({holders})"
    cur.execute(select_query, words)
    rows = cur.fetchall()
    con.commit()
    con.close()

    xlsx_write(rows, ['word', 'phonetic', 'translation'], out_file,
               'Words_Lemma_Sheet')


# init_ecdict_sqlite() # 只需运行一次，生成sqlite3 db文件
# write_from_words('FOO.txt')
