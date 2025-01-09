#!/usr/bin/env python3
'''通过每行一词或短语的文本文件生成含两个sheet的xlsx单词表

从生字表文件（每单词一行）获取单词短语，输出包含两个sheet的excel文件，一个sheet
含单词、词根、音标和中文释义。另一个sheet由前者词根生成,含单词，音标和中文释义。

Example:
    示例::

    if not Path(MYSQLITE).exists():
        init_ecdict_sqlite()  # 只需运行一次，生成sqlite3 db文件
    write_from_file('FOO.txt')

Note:

    别忘修改MYSQLITE变量

'''
import sqlite3
from pathlib import Path, PurePath

from openpyxl import Workbook, load_workbook

from ECDICT import stardict

MYSQLITE = 'ecdictSqlite.db'


def exchange2dict(orig_word, exchange):
    """ECDICT exchange字段的转为字典

    Note:
        ecdict exchange列中0: 代表Lemma，如 perceived 的 Lemma 是 perceive

        类型 	说明
        p 	过去式（did）
        d 	过去分词（done）
        i 	现在分词（doing）
        3 	第三人称单数（does）
        r 	形容词比较级（-er）
        t 	形容词最高级（-est）
        s 	名词复数形式
        0 	Lemma，如 perceived 的 Lemma 是 perceive
        1 	Lemma 的变换形式，比如 s 代表 apples 是其 lemma 的复数形式

        此外，还有f: b: z:等分段

        https://github.com/skywind3000/ECDICT/issues/23
    """

    lists = exchange.split('/')

    inflection = {}

    for trans in lists:
        if len(trans) < 3:
            continue
        inflection[trans[0]] = trans[2:]

    if not inflection.get('0'):
        inflection['0'] = orig_word

    return inflection


def init_ecdict_sqlite():
    """将ecdict.csv转换为sqlite数据库，并新增lemma列

    """
    stardict.convert_dict(MYSQLITE, 'ECDICT/ecdict.csv')

    con = sqlite3.connect(MYSQLITE)
    cur = con.cursor()
    cur.execute("ALTER TABLE stardict ADD lemma TEXT")

    # 查询所有行的数据
    select_query = "SELECT word, exchange FROM stardict"
    cur.execute(select_query)
    rows = cur.fetchall()

    # 遍历每一行，根据 word和exchange 的值更新新建列lemma
    for row in rows:
        word = row[0]
        exchange = row[1]

        lemma = exchange2dict(word, exchange)['0']

        # 更新 lemma 列的值
        update_query = "UPDATE stardict SET lemma = ? WHERE word = ?"
        cur.execute(update_query, (lemma, word))

    con.commit()
    con.close()


def xlsx_write(word_lines, columns, filename, sheetname='sheet1'):
    """将列表文件写入filename文件

    """

    if Path(filename).exists():
        wb = load_workbook(filename)
    else:
        wb = Workbook()

    # 如果sheetname已存在的话，openpyxl会依次在sheetname后追加1,2,3...
    worksheet = wb.create_sheet(sheetname)

    rows = len(word_lines)
    print('行数有：', rows)
    if rows <= 1:
        print(f'word_lines行数为{rows}，有误，不写入文件，退出')
        return

    cols = len(columns)
    if cols != len(word_lines[0]):
        print(
            '参数 columns 有 {cols}列，但word_lines第一行有{len(word_lines[0])}列\n列数匹配有误，程序退出'
        )

    for j in range(cols):
        worksheet.cell(1, j + 1, columns[j])

    # Iterate over the data and write it out row by row.
    for i in range(rows):
        for j in range(cols):
            worksheet.cell(i + 2, j + 1, word_lines[i][j])

    wb.save(filename)
    wb.close()


def write_from_words(words, out_file):
    """从单词表（每单词一行）获取单词，输出单词、音标和中文释义

    """
    con = sqlite3.connect(MYSQLITE)
    cur = con.cursor()

    # 查询所有行的数据
    holders = ', '.join('?' for _ in words)
    select_query = f"SELECT word, phonetic, translation FROM stardict WHERE word IN ({holders})"
    cur.execute(select_query, words)
    rows = cur.fetchall()
    con.commit()
    con.close()

    xlsx_write(rows, ['word', 'phonetic', 'translation'], out_file,
               'Words_Lemma_Sheet')


def write_from_file(input_file):
    """从生字表文件（每单词一行）获取单词短语，输出包含两个sheet的excel文件，一个
    sheet含单词、词根、音标和中文释义。另一个sheet由前者词根生成,含单词，音标和
    中文释义。

    """
    with open(input_file, 'r') as f:
        words = [line.strip() for line in f.readlines()]

    con = sqlite3.connect(MYSQLITE)
    cur = con.cursor()

    # 查询所有行的数据
    holders = ', '.join('?' for _ in words)
    select_query = f"SELECT word, lemma, phonetic, translation FROM stardict WHERE word IN ({holders})"
    cur.execute(select_query, words)
    rows = cur.fetchall()
    con.commit()
    con.close()

    out_file = PurePath(input_file).stem + '_dict.xlsx'

    xlsx_write(rows, ['word', 'lemma', 'phonetic', 'translation'], out_file,
               'Words_Sheet')

    lemmas = [j[1] for j in rows if j[1] is not None]
    write_from_words(lemmas, out_file)


if not Path(MYSQLITE).exists():
    init_ecdict_sqlite()  # 只需运行一次，生成sqlite3 db文件
write_from_file('1984_glossary.txt')
